"""Excel出力 (openpyxl)。"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _write_table(ws, rows, start_row=1):
    """dict listをシートに書き込む。ヘッダー付き。"""
    if not rows:
        ws.cell(row=start_row, column=1, value="リソースなし")
        return start_row + 1

    headers = list(rows[0].keys())
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP

    for row_idx, row in enumerate(rows, start_row + 1):
        for col_idx, h in enumerate(headers, 1):
            val = row.get(h, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val))
            cell.alignment = _WRAP

    # 列幅自動調整
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row in rows:
            val_len = len(str(row.get(h, "")))
            max_len = max(max_len, val_len)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    return start_row + len(rows) + 1


def _safe_sheet_name(name):
    """Excelシート名の制約(31文字、禁止文字)に対応。"""
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        name = name.replace(ch, "_")
    return name[:31]


def output_excel(inventory_data, estimate_flag, filepath):
    """inventory_dataをExcelファイルに出力。"""
    wb = Workbook()

    # --- サマリシート ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    meta = inventory_data["metadata"]
    summary = inventory_data["summary"]

    info = [
        ("Account", meta["account_id"]),
        ("Date", meta["date"]),
        ("Regions", ", ".join(meta["regions"])),
        ("Total Services", summary["total_services"]),
        ("Total Resources", summary["total_resources"]),
        ("Errors", len(summary["errors"])),
    ]
    if estimate_flag and inventory_data.get("total_monthly_usd") is not None:
        info.append(("Monthly Estimate (USD)", f"${inventory_data['total_monthly_usd']:,.2f}"))

    for row_idx, (label, val) in enumerate(info, 1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=str(val))
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 60

    # サービス別リソース数テーブル
    svc_start = len(info) + 2
    ws_summary.cell(row=svc_start, column=1, value="Service").font = _HEADER_FONT
    ws_summary.cell(row=svc_start, column=1).fill = _HEADER_FILL
    ws_summary.cell(row=svc_start, column=2, value="Count").font = _HEADER_FONT
    ws_summary.cell(row=svc_start, column=2).fill = _HEADER_FILL
    ws_summary.cell(row=svc_start, column=3, value="Warnings").font = _HEADER_FONT
    ws_summary.cell(row=svc_start, column=3).fill = _HEADER_FILL

    for i, svc in enumerate(inventory_data["services"], svc_start + 1):
        ws_summary.cell(row=i, column=1, value=svc["name"])
        ws_summary.cell(row=i, column=2, value=svc["count"])
        warn_count = len(svc.get("warnings", []))
        cell = ws_summary.cell(row=i, column=3, value=warn_count)
        if warn_count > 0:
            cell.fill = _WARN_FILL

    ws_summary.column_dimensions["C"].width = 12

    # --- サービスごとのシート ---
    for svc in inventory_data["services"]:
        sheet_name = _safe_sheet_name(svc["name"])
        ws = wb.create_sheet(title=sheet_name)
        next_row = _write_table(ws, svc["resources"])

        # 概算があれば同シートに追記
        if svc.get("estimates"):
            next_row += 1
            ws.cell(row=next_row, column=1, value="月額概算").font = Font(bold=True, size=12)
            next_row += 1
            _write_table(ws, svc["estimates"], start_row=next_row)

    # --- 警告シート ---
    all_warnings = []
    for svc in inventory_data["services"]:
        for w in svc.get("warnings", []):
            all_warnings.append({"Service": svc["name"], **w})

    if all_warnings:
        ws_warn = wb.create_sheet(title="⚠ Warnings")
        _write_table(ws_warn, all_warnings)
        # 警告行を黄色ハイライト
        for row_idx in range(2, len(all_warnings) + 2):
            for col_idx in range(1, len(all_warnings[0]) + 1):
                ws_warn.cell(row=row_idx, column=col_idx).fill = _WARN_FILL

    # --- 概算合計シート ---
    if estimate_flag:
        all_estimates = []
        for svc in inventory_data["services"]:
            for e in svc.get("estimates", []):
                all_estimates.append({"Service": svc["name"], **e})

        if all_estimates:
            ws_est = wb.create_sheet(title="Cost Estimate")
            next_row = _write_table(ws_est, all_estimates)
            next_row += 1
            ws_est.cell(row=next_row, column=1, value="合計").font = Font(bold=True)
            if inventory_data.get("total_monthly_usd") is not None:
                ws_est.cell(row=next_row, column=len(all_estimates[0]),
                            value=f"${inventory_data['total_monthly_usd']:,.2f}").font = Font(bold=True)

    # エラーシート
    if summary["errors"]:
        ws_err = wb.create_sheet(title="Errors")
        ws_err.cell(row=1, column=1, value="Error").font = _HEADER_FONT
        ws_err.cell(row=1, column=1).fill = _HEADER_FILL
        for i, err in enumerate(summary["errors"], 2):
            ws_err.cell(row=i, column=1, value=err)
        ws_err.column_dimensions["A"].width = 100

    wb.save(filepath)
